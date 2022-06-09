from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import sys

def multiply_excel_table(limit):
    # assign the workbook and worksheet variables
    wb = load_workbook('multiplication_table.xlsx')
    ws = wb.active

    # insert the multiplicants to the table
    for i in range(2, limit + 2):
        ws['A' + str(i)] = i - 1
        ws[get_column_letter(i) + '1'] = i - 1
        # make them bold
        ws['A' + str(i)].font = Font(bold=True)
        ws[get_column_letter(i) + '1'].font = Font(bold=True)

    # assign multiplication values
    for row in range(2, limit + 2):
        for col in range(2, limit + 2):
            # convert the number to the column character
            col_char = get_column_letter(col)
            # do the multiplication
            ws[col_char + str(row)] = ws['A' + str(row)].value * ws[col_char + '1'].value

    # save the file
    wb.save('multiplication_table.xlsx')

def main():
    # get the limit of the multiplication table from the user
    limit = int(sys.argv[1])
    multiply_excel_table(limit)
    sys.exit()

if __name__ == '__main__':
    main()
