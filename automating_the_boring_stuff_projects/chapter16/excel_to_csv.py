import os
import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def excel_to_csv():
    for excelFile in os.listdir('.\excel_files'):
        # Skip non-xlsx files, load the workbook object.
        wb = load_workbook('.\excel_files\\' + excelFile)

        for sheetName in wb.sheetnames:
            # Loop through every sheet in the workbook.
            for sheetName in wb.sheetnames:
                sheet = wb[sheetName]

                # Create the CSV filename from the Excel filename and sheet title.
                csv_file = open(excelFile[:len(excelFile) - 5] + '_' + sheet.title + '.csv', 'w', newline='')
                # Create the csv.writer object for this CSV file.
                csv_writer = csv.writer(csv_file)

            # Loop through every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                rowData = [] # append each cell to this list
                # Loop through each cell in the row.
                for colNum in range(1, sheet.max_column + 1):
                    rowData.append(sheet[get_column_letter(colNum) + str(rowNum)].value)
                # Write the rowData list to the CSV file.
                csv_writer.writerow(rowData)
            csv_file.close()

def main():
    excel_to_csv()

if __name__ == '__main__':
    main()
