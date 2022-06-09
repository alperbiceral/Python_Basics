from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def invert_cells():
    # load the workbook and worksheet
    wb = load_workbook('cell_inverted.xlsx')
    ws = wb.active

    col_num = 2
    row_num = 3

    # create the lists for reading columns
    cols = [['ITEM:'], ['SOLD:']]

    # read the data from the file into a list
    for col in range(1, col_num + 1):
        col_char = get_column_letter(col)
        for row in range(1, row_num + 1):
           cols[col - 1].append(ws[col_char + str(row)].value)

    # delete columns
    for col in range(1, col_num + 1):
        ws.delete_cols(1)

    # insert the numbers to the file
    for row in range(1, col_num + 1):
            ws.append(cols[row - 1])

    # save the file
    wb.save('cell_inverted.xlsx')

def main():
    invert_cells()

if __name__ == '__main__':
    main()
